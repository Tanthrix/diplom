from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import CustomUser, Device, Specification, Role, UserRole, Room
from .forms import CustomUserCreationForm, CustomUserEditForm, DeviceForm, SpecificationForm, RoleForm, RoomForm
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
import openpyxl
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages

@login_required
def admin_dashboard(request):
    return render(request, 'disbursement/dashboard.html')

@login_required
def logout_view(request):
    from django.contrib.auth import logout
    logout(request)
    return redirect('login')

@login_required
def user_management(request):
    users = CustomUser.objects.prefetch_related('user_roles__role').all()
    return render(request, 'disbursement/users.html', {'users': users})

@login_required
def add_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            for role in form.cleaned_data['roles']:
                UserRole.objects.create(user=user, role=role)
            messages.success(request, 'Пользователь успешно создан.')
            return redirect('user_management')
    else:
        form = CustomUserCreationForm()
    return render(request, 'disbursement/add_user.html', {'form': form})

@login_required
def edit_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    if request.method == 'POST':
        form = CustomUserEditForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save()
            current_roles = set(user.user_roles.all().values_list('role_id', flat=True))
            new_roles = set(form.cleaned_data['roles'].values_list('id', flat=True))
            roles_to_remove = current_roles - new_roles
            if roles_to_remove:
                UserRole.objects.filter(user=user, role_id__in=roles_to_remove).delete()
            roles_to_add = new_roles - current_roles
            for role_id in roles_to_add:
                UserRole.objects.create(user=user, role_id=role_id)
            messages.success(request, 'Пользователь обновлён.')
            return redirect('user_management')
    else:
        form = CustomUserEditForm(instance=user, initial={'roles': [ur.role for ur in user.user_roles.all()]})
    return render(request, 'disbursement/edit_user.html', {'form': form, 'user': user})

@login_required
def delete_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    user.delete()
    messages.success(request, 'Пользователь удалён.')
    return redirect('user_management')

@login_required
def device_list(request):
    devices = Device.objects.select_related('added_by', 'room').prefetch_related('specifications').all()
    return render(request, 'disbursement/device_list.html', {'devices': devices})

@login_required
def device_create(request):
    if request.method == 'POST':
        device_form = DeviceForm(request.POST, request.FILES)
        if device_form.is_valid():
            device = device_form.save(commit=False)
            device.added_by = request.user
            device.save()
            messages.success(request, 'Устройство добавлено.')
            return redirect('device_list')
    else:
        device_form = DeviceForm()
    return render(request, 'disbursement/device_form.html', {
        'device_form': device_form
    })

@login_required
def edit_device(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    if request.method == 'POST':
        device_form = DeviceForm(request.POST, request.FILES, instance=device)
        if device_form.is_valid():
            device_form.save()
            messages.success(request, 'Устройство обновлено.')
            return redirect('device_list')
    else:
        device_form = DeviceForm(instance=device)
    return render(request, 'disbursement/device_form.html', {
        'device_form': device_form,
        'device': device
    })

@login_required
def delete_device(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    device.delete()
    messages.success(request, 'Устройство удалено.')
    return redirect('device_list')

@login_required
def add_specification(request, device_id):
    device = get_object_or_404(Device, id=device_id)
    if request.method == 'POST':
        form = SpecificationForm(request.POST)
        if form.is_valid():
            specification = form.save(commit=False)
            specification.device = device
            specification.save()
            messages.success(request, 'Характеристика добавлена.')
            return redirect('device_list')
    else:
        form = SpecificationForm()
    return render(request, 'disbursement/add_specification.html', {
        'form': form,
        'device': device
    })

@login_required
def edit_specification(request, spec_id):
    specification = get_object_or_404(Specification, id=spec_id)
    if request.method == 'POST':
        form = SpecificationForm(request.POST, instance=specification)
        if form.is_valid():
            form.save()
            messages.success(request, 'Характеристика обновлена.')
            return redirect('device_list')
    else:
        form = SpecificationForm(instance=specification)
    return render(request, 'disbursement/edit_specification.html', {
        'form': form,
        'specification': specification
    })

@login_required
def delete_specification(request, spec_id):
    specification = get_object_or_404(Specification, id=spec_id)
    specification.delete()
    messages.success(request, 'Характеристика удалена.')
    return redirect('device_list')

@login_required
def import_export_devices(request):
    if request.method == 'POST':
        if 'export_excel' in request.POST:
            wb = Workbook()
            ws = wb.active
            ws.title = "Устройства"

            headers = ['ID', 'Название', 'Серийный номер', 'Статус', 'Аудитория', 'Добавлено', 'Добавил', 'Характеристики']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)

            devices = Device.objects.select_related('added_by', 'room').prefetch_related('specifications').all()
            row = 2
            for device in devices:
                specs = "; ".join([f"{spec.name}: {spec.value}" for spec in device.specifications.all()])
                ws.cell(row=row, column=1).value = device.id
                ws.cell(row=row, column=2).value = device.name
                ws.cell(row=row, column=3).value = device.serial_number or "Не указан"
                ws.cell(row=row, column=4).value = device.get_status_display()
                ws.cell(row=row, column=5).value = device.room.number if device.room else "Не указана"
                ws.cell(row=row, column=6).value = device.created_at.strftime("%d.%m.%Y %H:%M")
                ws.cell(row=row, column=7).value = device.added_by.full_name
                ws.cell(row=row, column=8).value = specs
                row += 1

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="devices.xlsx"'
            return response

        elif 'import_excel' in request.POST and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
            path = default_storage.save('tmp/import.xlsx', ContentFile(excel_file.read()))
            wb = openpyxl.load_workbook(default_storage.path(path))
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                device_id, name, serial_number, status, room_number, created_at, added_by, specs = row
                if not name:
                    continue

                try:
                    user = CustomUser.objects.get(full_name=added_by)
                except CustomUser.DoesNotExist:
                    messages.error(request, f'Пользователь {added_by} не найден.')
                    continue

                room = None
                if room_number and room_number != "Не указана":
                    room, _ = Room.objects.get_or_create(number=room_number)

                status_map = {
                    'Доступно': 'available',
                    'В использовании': 'in_use',
                    'На ремонте': 'under_repair',
                }
                status = status_map.get(status, 'available')

                device, created = Device.objects.update_or_create(
                    id=device_id if device_id else None,
                    defaults={
                        'name': name,
                        'serial_number': serial_number,
                        'status': status,
                        'added_by': user,
                        'room': room,
                    }
                )

                if specs:
                    device.specifications.all().delete()
                    for spec in specs.split(";"):
                        spec = spec.strip()
                        if ":" in spec:
                            name, value = spec.split(":", 1)
                            Specification.objects.create(device=device, name=name.strip(), value=value.strip())

            default_storage.delete(path)
            messages.success(request, 'Устройства импортированы.')
            return redirect('import_export_devices')

    return render(request, 'disbursement/import_export_devices.html')

@login_required
def role_list(request):
    roles = Role.objects.all()
    return render(request, 'disbursement/role_list.html', {'roles': roles})

@login_required
def add_role(request):
    if request.method == 'POST':
        form = RoleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Роль добавлена.')
            return redirect('role_list')
    else:
        form = RoleForm()
    return render(request, 'disbursement/add_role.html', {'form': form})

@login_required
def edit_role(request, role_id):
    role = get_object_or_404(Role, id=role_id)
    if request.method == 'POST':
        form = RoleForm(request.POST, instance=role)
        if form.is_valid():
            form.save()
            messages.success(request, 'Роль обновлена.')
            return redirect('role_list')
    else:
        form = RoleForm(instance=role)
    return render(request, 'disbursement/edit_role.html', {'form': form, 'role': role})

@login_required
def delete_role(request, role_id):
    role = get_object_or_404(Role, id=role_id)
    if request.method == 'POST':
        role.delete()
        messages.success(request, 'Роль удалена.')
        return redirect('role_list')
    return render(request, 'disbursement/confirm_delete.html', {'object': role, 'type': 'роль'})

@login_required
def user_role_list(request):
    user_roles = UserRole.objects.select_related('user', 'role').all()
    return render(request, 'disbursement/user_role_list.html', {'user_roles': user_roles})

@login_required
def delete_user_role(request, user_role_id):
    user_role = get_object_or_404(UserRole, id=user_role_id)
    if request.method == 'POST':
        user_role.delete()
        messages.success(request, 'Связь удалена.')
        return redirect('user_role_list')
    return render(request, 'disbursement/confirm_delete.html', {'object': user_role, 'type': 'связь пользователя и роли'})

@login_required
def room_list(request):
    rooms = Room.objects.prefetch_related('devices').all()
    return render(request, 'disbursement/room_list.html', {'rooms': rooms})

@login_required
def add_room(request):
    if request.method == 'POST':
        form = RoomForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Аудитория добавлена.')
            return redirect('room_list')
    else:
        form = RoomForm()
    return render(request, 'disbursement/add_room.html', {'form': form})

@login_required
def room_detail(request, room_id):
    room = get_object_or_404(Room, id=room_id)
    devices = room.devices.all()
    return render(request, 'disbursement/room_detail.html', {'room': room, 'devices': devices})

def custom_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Вы не являетесь администратором или неверные логин/пароль.')
    return render(request, 'disbursement/login.html')